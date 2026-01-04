import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side


class ExcelUtils:
    """
    Excel工具类
    """

    def __init__(self, json_path: Path, excel_path: Path) -> None:
        """
        初始化
        :param json_path: json文件路径
        :param excel_path: excel文件路径
        """
        self.json_path = json_path
        self.excel_path = excel_path

    def json_to_excel(self) -> None:
        """
        将json文件转换为excel文件
        :return: None
        """
        try:
            with open(self.json_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                for key in data:
                    df = pd.DataFrame(data[key])
                    df.to_excel(writer, sheet_name=key, index=False)
        except FileNotFoundError:
            print(f"错误：找不到JSON文件 {self.json_path}")
        except json.JSONDecodeError:
            print(f"错误：JSON文件格式不正确 {self.json_path}")
        except Exception as e:
            print(f"转换JSON到Excel时发生错误: {e}")

    @staticmethod
    def _col_index_to_letter(index) -> str:
        """
        将列索引转换为列字母
        :param index: 列索引
        :return: 列字母
        """
        if index <= 0:
            return ""

        letter: str = ""
        while index > 0:
            index -= 1
            letter = chr(index % 26 + ord("A")) + letter
            index = index // 26
        return letter

    @staticmethod
    def _find_special_columns(ws) -> tuple:
        """
        查找PC和移动列的索引
        :param ws: 工作表对象
        :return: (pc_col, mobile_col) 元组
        """
        pc_col = None
        mobile_col = None

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value == "PC":
                pc_col = col_idx
            elif cell_value == "移动":
                mobile_col = col_idx

        return pc_col, mobile_col

    def set_wrap_text(self, wb: Workbook, sheet_name: str) -> None:
        """
        设置PC和移动列的换行和宽度
        :param wb: 工作簿
        :param sheet_name: 工作表名称
        :return: None
        """
        try:
            ws = wb[sheet_name]
            pc_col, mobile_col = self._find_special_columns(ws)

            # 设置PC列的格式
            if pc_col:
                self._set_column_format(ws, pc_col, 28)

            # 设置移动列的格式
            if mobile_col:
                self._set_column_format(ws, mobile_col, 28)

        except Exception as e:
            print(f"设置换行格式时出错: {e}")

    def _set_column_format(self, ws, col_index: int, width: int) -> None:
        """
        设置列的宽度和对齐方式
        :param ws: 工作表
        :param col_index: 列索引
        :param width: 列宽
        """
        col_letter = self._col_index_to_letter(col_index)
        ws.column_dimensions[col_letter].width = width

        for cell in ws.iter_cols(min_col=col_index, max_col=col_index, min_row=1):
            for c in cell:
                c.alignment = Alignment(wrapText=True)

    @staticmethod
    def set_row_height(wb: Workbook, sheet_name: str) -> None:
        """
        设置行高
        :param wb: 工作簿
        :param sheet_name: 工作表名称
        :return: None
        """
        try:
            ws = wb[sheet_name]
            for row_num in range(2, ws.max_row + 1):
                ws.row_dimensions[row_num].height = 78
                for cell in ws[row_num]:
                    cell.alignment = Alignment(vertical="center")
        except Exception as e:
            print(f"设置行高时出错: {e}")

    def set_row_width(self, wb: Workbook, sheet_name: str) -> None:
        """
        设置列宽（除PC和移动列外）
        :param wb: 工作簿
        :param sheet_name: 工作表名称
        :return: None
        """
        try:
            ws = wb[sheet_name]
            pc_col, mobile_col = self._find_special_columns(ws)

            for col_idx in range(1, ws.max_column + 1):
                if col_idx == pc_col or col_idx == mobile_col:
                    continue  # 跳过PC和移动列

                max_length = self._calculate_max_column_length(ws, col_idx)
                col_letter = self._col_index_to_letter(col_idx)
                adjusted_width = max(max_length + 2, 10)  # 最小宽度为10
                ws.column_dimensions[col_letter].width = adjusted_width

        except Exception as e:
            print(f"设置列宽时出错: {e}")

    @staticmethod
    def _calculate_max_column_length(ws, col_index: int) -> int:
        """
        计算列中最长内容的长度
        :param ws: 工作表
        :param col_index: 列索引
        :return: 最大长度
        """
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_index).value
            if cell_value:
                try:
                    # 计算单元格内容长度，中文算2个字符，英文和数字算1个
                    content = str(cell_value)
                    length = sum(2 if ord(c) > 127 else 1 for c in content)
                    if length > max_length:
                        max_length = length
                except Exception:
                    # 如果转换失败，跳过该单元格
                    continue
        return max_length

    @staticmethod
    def set_border(wb: Workbook) -> None:
        """
        设置边框
        :param wb: 工作簿
        :return: None
        """
        # 定义边框样式
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 遍历所有有数据的单元格并设置边框
                for row in ws.iter_rows(
                    min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                ):
                    for cell in row:
                        cell.border = thin_border
        except Exception as e:
            print(f"设置边框时出错: {e}")

    def set_format(self) -> None:
        """
        设置格式
        :return: None
        """
        try:
            wb = load_workbook(self.excel_path)
            for sheet_name in wb.sheetnames:
                self.set_wrap_text(wb, sheet_name)
                self.set_row_height(wb, sheet_name)
                self.set_row_width(wb, sheet_name)
                self.set_border(wb)

            wb.save(self.excel_path)
        except Exception as e:
            print(f"设置格式时出错: {e}")
